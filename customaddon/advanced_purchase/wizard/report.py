import datetime
import base64
import io
from openpyxl import load_workbook
from odoo.modules import get_module_resource
import openpyxl
from copy import copy
from odoo.exceptions import UserError
from odoo.tools.translate import _
from odoo import fields, models, api


class ModelName(models.Model):
    _name = 'report.spending'

    month = fields.Selection([
        ('1', 1), ('2', 2), ('3', 3), ('4', 4),
        ('5', 5), ('6', 6), ('7', 7), ('8', 8),
        ('9', 9), ('10', 10), ('11', 11),
        ('12', 12), ], string="Tháng: ", default=str(fields.datetime.now().month),
        required=True
    )
    department = fields.Many2many("hr.department")

    def action_export_data(self):
        self.ensure_one()

        wb = load_workbook(get_module_resource('advanced_purchase', 'static/template',
                                               'bao_cao_chi_tieu.xlsx'))
        ws = wb['Sheet1']
        style = copy(ws['A8']._style)

        month = int(self.month)
        current_time = datetime.datetime.today().replace(day=1, month=month)  # 3/2/2023
        date_from = current_time
        date_to = datetime.datetime.today().replace(day=1, month=month + 1)
        # tao ban ghi model moi

        if self.department:
            po = self.env['purchase.order'].sudo().search(
                [('department', 'in', self.department.ids), ('state', '=', 'purchase'),
                 ('create_date', '>=', date_from), ('create_date', '<', date_to)])
            dict_data = []
            for dep in self.department.ids:
                self.env['hr.department'].browse(dep).cost = 0
            for order in po:
                for dep in self.department.ids:
                    if dep == order.department.id:
                        self.env['hr.department'].browse(dep).cost = self.env['hr.department'].browse(
                            dep).cost + order.amount_total

            for dep in self.department.ids:
                dict_data.append({"department_id": dep,
                                  "cost": self.env['hr.department'].browse(dep).cost,
                                  "spending_limit": self.env['hr.department'].browse(dep).spending_limit_a_month})
            if dict_data:
                row = 8
                stt = 1
                for line in dict_data:
                    ws.cell(row=row, column=1).value = stt
                    ws.cell(row=row, column=1)._style = style
                    ws.cell(row=row, column=2).value = line['department_id']
                    ws.cell(row=row, column=2)._style = style
                    ws.cell(row=row, column=3).value = line['cost']
                    ws.cell(row=row, column=3)._style = style
                    ws.cell(row=row, column=4).value = line['spending_limit']
                    ws.cell(row=row, column=4)._style = style
                    stt += 1
                    row += 1

                # ws.cell(row=4, column=6).value = date
                ws.cell(row=5, column=2).value = 'BÁO CÁO MUA HÀNG THEO THÁNG ' + self.month
                content = io.BytesIO()
                # wb.save(content)
                wb.save(content)
                out = base64.encodebytes(content.getvalue())

                file_output = self.env['result.export.report'].sudo().create({
                    'file': out,
                    'file_name': "bao_cao_chi_tieu.xlsx"
                })
                content.close()
                return {
                    'type': 'ir.actions.act_url',
                    'target': 'new',
                    'url': 'web/content/?model=' + file_output._name + '&id=' + str(
                        file_output.id) + '&field=file&download=true&filename=' + str(file_output.file_name),
                }

            else:
                raise UserError(_("Không có dữ liệu để xuất."))

            # delete all recort

            self.env['report'].search([(1, '=', 1)]).unlink()

            # create new records
            for x in dict_data:
                self.env['report'].create({
                    "department": x['department_id'],
                    "cost": x['cost'],
                    "spending_limit": x['spending_limit'],
                })

            return {
                "name": "Department",
                'res_model': 'report',
                'type': 'ir.actions.act_window',
                'view_mode': 'tree',
                'view_id': self.env.ref("advanced_purchase.report_tree_view").id,
                'target': 'current',
                # 'domain': [('state', '=', 'purchase'),('create_date', '>=', date_from),('create_date', '<', date_to), ('department', 'in', self.department.ids)],
            }

        else:
            po = self.env['purchase.order'].sudo().search(
                [('state', '=', 'purchase'), ('create_date', '>=', date_from), ('create_date', '<', date_to)])
            dict_report = []
            list_dep = []
            for dep in self.env['hr.department'].search([]):
                dep.cost = 0.0

            for order in po:
                if order.department.id not in list_dep:
                    for dep in self.env['hr.department'].search([]):
                        if dep.id == order.department.id:
                            dep.cost = dep.cost + order.amount_total

            for dep in self.env['hr.department'].search([]):
                dict_report.append(
                    {"department_id": dep.id, "cost": dep.cost, "spending_limit": dep.spending_limit_a_month})

            self.env['report'].search([]).unlink()

            for x in dict_report:
                self.env['report'].create({
                    "department": x['department_id'],
                    "cost": x['cost'],
                    "spending_limit": x['spending_limit'],
                })

            return {
                "name": "Department",
                'res_model': 'report',
                'type': 'ir.actions.act_window',
                'view_mode': 'tree',
                'view_id': self.env.ref("advanced_purchase.report_tree_view").id,
                'target': 'current',
            }
